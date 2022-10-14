from datetime import datetime
from typing import List, Dict, Any
from os.path import exists
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.base import pred_type, get_items


def load_or_create_wb(name: str) -> openpyxl.Workbook:
    if exists(name):
        wb = openpyxl.load_workbook(name)
    else:
        wb = openpyxl.Workbook()
    return wb


def get_or_create_sheet(wb: openpyxl.Workbook, model_dict: Dict[str, Any]) -> Worksheet:
    _, name, temp_dict = get_items(model_dict)
    arguments = list(temp_dict.keys())
    if name in wb.sheetnames:
        return wb[name]

    ws = wb.create_sheet(title=name)
    ws.title = name

    beginning = ["Date", "MAPE"]
    ending = [f"+{x}" for x in range(1, 6)]
    headings = beginning + arguments + ending
    for i, column in enumerate(headings):
        ws.cell(row=1, column=i + 1, value=column)
    return ws


def get_empty(ws: Worksheet) -> int:
    first_blank = 1
    while ws[f"A{first_blank}"].value is not None:
        first_blank += 1
    return first_blank


def enter_row(
    ws: Worksheet, row: int, date: datetime, data: pred_type, model_dict: Dict[str, Any]
) -> None:
    _, _, temp_dict = get_items(model_dict)
    beginning: List[Any] = [date, data[1] / 100]
    ending: List[Any] = data[0]
    values: List[Any] = beginning + list(temp_dict.values()) + ending
    for i, val in enumerate(values):
        ws.cell(row=row, column=i + 1, value=val)


def existing_dates(ws) -> List[str]:
    dates = []
    for i in range(2, ws.max_row + 1):
        if ws[f"A{i}"].value is not None:
            value = ws[f"A{i}"].value
            dates.append(value)
    return dates
